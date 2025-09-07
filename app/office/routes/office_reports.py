from app.models import (
    Inquiry, InquiryMessage, User, Office, db, OfficeAdmin, 
    Student, CounselingSession, StudentActivityLog, SuperAdminActivityLog, 
    OfficeLoginLog, AuditLog, Announcement, ConcernType, OfficeConcernType
)
from flask import Blueprint, redirect, url_for, render_template, jsonify, request, flash, Response, send_file, current_app
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from sqlalchemy import func, case, desc, or_, and_
from app.office import office_bp
import io
import csv
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def get_logo_path(office=None):
    """Return the most appropriate logo path.

    Fallback chain:
    1. Office-specific branding folder: static/uploads/branding/<office_id>.png|jpg
    2. Default school logo in static/images/schoollogo.png
    3. Returns None if nothing exists.
    """
    try:
        base = current_app.root_path
        # Office specific
        if office and getattr(office, 'id', None):
            branding_dir = os.path.join(base, 'static', 'uploads', 'branding')
            for ext in ('png', 'jpg', 'jpeg', 'gif'):
                candidate = os.path.join(branding_dir, f"{office.id}.{ext}")
                if os.path.exists(candidate):
                    return candidate
        # Default
        default_logo = os.path.join(base, 'static', 'images', 'schoollogo.png')
        if os.path.exists(default_logo):
            return default_logo
    except Exception:
        pass
    return None


@office_bp.route('/reports')
@login_required
def office_reports():
    """Main reports dashboard page"""
    office_id = current_user.office_admin.office_id
    
    # Get report statistics
    stats = get_report_stats(office_id)
    
    # Get concern types that are supported by this office only
    concern_types = db.session.query(ConcernType)\
        .join(OfficeConcernType)\
        .filter(OfficeConcernType.office_id == office_id)\
        .all()
    
    return render_template('office/office_reports.html', 
                         stats=stats, 
                         concern_types=concern_types)


@office_bp.route('/reports/generate', methods=['POST'])
@login_required
def generate_report():
    """Generate and download reports based on user selection"""
    office_id = current_user.office_admin.office_id
    
    data = request.get_json()
    
    # Debug logging
    print(f"Received data: {data}")
    
    if not data:
        return jsonify({'error': 'No data received'}), 400
    
    report_type = data.get('report_type')
    date_range = data.get('date_range')
    format_type = data.get('format')
    filters = data.get('filters', {})
    
    # Validation
    if not report_type:
        return jsonify({'error': 'Report type is required'}), 400
    
    if not date_range:
        return jsonify({'error': 'Date range is required'}), 400
        
    if not format_type:
        return jsonify({'error': 'Export format is required'}), 400
    
    print(f"Report type: {report_type}, Date range: {date_range}, Format: {format_type}")
    
    # Parse date range
    start_date, end_date = parse_date_range(date_range)
    
    try:
        if report_type == 'inquiries':
            return generate_inquiries_report(office_id, start_date, end_date, format_type, filters)
        elif report_type == 'counseling':
            return generate_counseling_report(office_id, start_date, end_date, format_type, filters)
        elif report_type == 'activity':
            return generate_activity_report(office_id, start_date, end_date, format_type, filters)
        elif report_type == 'summary':
            return generate_summary_report(office_id, start_date, end_date, format_type, filters)
        else:
            return jsonify({'error': f'Invalid report type: {report_type}'}), 400
            
    except Exception as e:
        import traceback
        print(f"Error generating report: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500


def get_report_stats(office_id):
    """Get statistics for the reports dashboard"""
    now = datetime.utcnow()
    thirty_days_ago = now - timedelta(days=30)
    
    # Total inquiries
    total_inquiries = Inquiry.query.filter_by(office_id=office_id).count()
    
    # Inquiries this month
    inquiries_this_month = Inquiry.query.filter(
        Inquiry.office_id == office_id,
        Inquiry.created_at >= thirty_days_ago
    ).count()
    
    # Total counseling sessions
    total_sessions = CounselingSession.query.filter_by(office_id=office_id).count()
    
    # Sessions this month
    sessions_this_month = CounselingSession.query.filter(
        CounselingSession.office_id == office_id,
        CounselingSession.scheduled_at >= thirty_days_ago
    ).count()
    
    # Response rate calculation
    responded_inquiries = Inquiry.query.filter(
        Inquiry.office_id == office_id,
        Inquiry.status.in_(['resolved', 'in_progress'])
    ).count()
    
    response_rate = (responded_inquiries / total_inquiries * 100) if total_inquiries > 0 else 0
    
    # Average resolution time (simplified)
    avg_resolution_time = "2.5 days"  # This would need proper calculation
    
    return {
        'total_inquiries': total_inquiries,
        'inquiries_this_month': inquiries_this_month,
        'total_sessions': total_sessions,
        'sessions_this_month': sessions_this_month,
        'response_rate': round(response_rate, 1),
        'avg_resolution_time': avg_resolution_time
    }


def parse_date_range(date_range):
    """Parse date range string into start and end dates"""
    now = datetime.utcnow()
    
    if date_range == 'last_7_days':
        start_date = now - timedelta(days=7)
        end_date = now
    elif date_range == 'last_30_days':
        start_date = now - timedelta(days=30)
        end_date = now
    elif date_range == 'last_90_days':
        start_date = now - timedelta(days=90)
        end_date = now
    elif date_range == 'current_month':
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif date_range == 'current_year':
        start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    else:
        # Default to last 30 days
        start_date = now - timedelta(days=30)
        end_date = now
    
    return start_date, end_date


def generate_inquiries_report(office_id, start_date, end_date, format_type, filters):
    """Generate inquiries report"""
    # Build query
    query = Inquiry.query.filter(
        Inquiry.office_id == office_id,
        Inquiry.created_at >= start_date,
        Inquiry.created_at <= end_date
    )
    
    # Apply filters
    if filters.get('status'):
        query = query.filter(Inquiry.status == filters['status'])
    
    if filters.get('concern_type'):
        query = query.join(Inquiry.concerns).filter(
            Inquiry.concerns.any(concern_type_id=filters['concern_type'])
        )
    
    inquiries = query.order_by(desc(Inquiry.created_at)).all()
    
    if format_type == 'csv':
        return generate_csv_response(inquiries, 'inquiries')
    elif format_type == 'excel':
        return generate_excel_response(inquiries, 'inquiries')
    elif format_type == 'pdf':
        return generate_pdf_response(inquiries, 'inquiries', start_date, end_date)
    
    return jsonify({'error': 'Invalid format type'}), 400


def generate_counseling_report(office_id, start_date, end_date, format_type, filters):
    """Generate counseling sessions report"""
    query = CounselingSession.query.filter(
        CounselingSession.office_id == office_id,
        CounselingSession.scheduled_at >= start_date,
        CounselingSession.scheduled_at <= end_date
    )
    
    # Apply filters
    if filters.get('status'):
        query = query.filter(CounselingSession.status == filters['status'])
    
    sessions = query.order_by(desc(CounselingSession.scheduled_at)).all()
    
    if format_type == 'csv':
        return generate_csv_response(sessions, 'counseling')
    elif format_type == 'excel':
        return generate_excel_response(sessions, 'counseling')
    elif format_type == 'pdf':
        return generate_pdf_response(sessions, 'counseling', start_date, end_date)
    
    return jsonify({'error': 'Invalid format type'}), 400


def generate_activity_report(office_id, start_date, end_date, format_type, filters):
    """Generate activity logs report"""
    # This would query activity logs if available
    # For now, return empty data
    return jsonify({'message': 'Activity report generation not yet implemented'}), 501


def generate_summary_report(office_id, start_date, end_date, format_type, filters):
    """Generate summary report with statistics"""
    stats = get_detailed_stats(office_id, start_date, end_date)
    
    if format_type == 'pdf':
        return generate_summary_pdf(stats, office_id, start_date, end_date)
    elif format_type == 'excel':
        return generate_summary_excel(stats)
    
    return jsonify({'error': 'Invalid format type for summary report'}), 400


def get_detailed_stats(office_id, start_date, end_date):
    """Get detailed statistics for the given period"""
    # Implementation would include detailed statistics
    # For now, return basic stats
    return get_report_stats(office_id)


def generate_csv_response(data, report_type):
    """Generate CSV response"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    if report_type == 'inquiries':
        # Write CSV headers
        writer.writerow(['ID', 'Subject', 'Student Name', 'Student Email', 'Status', 'Created Date', 'Concerns'])
        
        # Write data rows
        for inquiry in data:
            concerns = ', '.join([concern.concern_type.name for concern in inquiry.concerns])
            writer.writerow([
                inquiry.id,
                inquiry.subject,
                inquiry.student.user.get_full_name(),
                inquiry.student.user.email,
                inquiry.status,
                inquiry.created_at.strftime('%Y-%m-%d %H:%M'),
                concerns
            ])
    
    elif report_type == 'counseling':
        writer.writerow(['ID', 'Student Name', 'Student Email', 'Status', 'Scheduled Date', 'Duration', 'Type'])
        
        for session in data:
            writer.writerow([
                session.id,
                session.student.user.get_full_name(),
                session.student.user.email,
                session.status,
                session.scheduled_at.strftime('%Y-%m-%d %H:%M') if session.scheduled_at else '',
                f"{session.duration_minutes} minutes" if session.duration_minutes else '',
                'Video' if session.is_video_session else 'In-person'
            ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={report_type}_report.csv'}
    )


def generate_excel_response(data, report_type):
    """Generate Excel response"""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"{report_type.title()} Report"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if report_type == 'inquiries':
        # Headers
        headers = ['ID', 'Subject', 'Student Name', 'Student Email', 'Status', 'Created Date', 'Concerns']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, inquiry in enumerate(data, 2):
            concerns = ', '.join([concern.concern_type.name for concern in inquiry.concerns])
            row_data = [
                inquiry.id,
                inquiry.subject,
                inquiry.student.user.get_full_name(),
                inquiry.student.user.email,
                inquiry.status,
                inquiry.created_at.strftime('%Y-%m-%d %H:%M'),
                concerns
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
    
    elif report_type == 'counseling':
        headers = ['ID', 'Student Name', 'Student Email', 'Status', 'Scheduled Date', 'Duration', 'Type']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row, session in enumerate(data, 2):
            row_data = [
                session.id,
                session.student.user.get_full_name(),
                session.student.user.email,
                session.status,
                session.scheduled_at.strftime('%Y-%m-%d %H:%M') if session.scheduled_at else '',
                f"{session.duration_minutes} minutes" if session.duration_minutes else '',
                'Video' if session.is_video_session else 'In-person'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={report_type}_report.xlsx'}
    )


def generate_pdf_response(data, report_type, start_date, end_date):
    """Generate PDF response with modern, professional styling.

    Enhancements:
    - Unified header & footer with logo, title, timestamp, page numbers
    - Clear sections (Summary / Details)
    - Balanced spacing & consistent font sizing
    - Status color badges (cell background colors)
    - Compact metric cards layout instead of plain key/value list
    - Graceful handling of empty datasets
    """
    output = io.BytesIO()

    # Capture generation timestamp once
    generated_at = datetime.now()
    generated_at_text = generated_at.strftime('%B %d, %Y • %I:%M %p')

    # Build dynamic period description
    if data:
        try:
            # Derive actual min/max from records if possible
            if report_type == 'inquiries':
                dates = [i.created_at for i in data if getattr(i, 'created_at', None)]
            else:
                dates = [s.scheduled_at for s in data if getattr(s, 'scheduled_at', None)]
            if dates:
                actual_start = min(dates).strftime('%b %d, %Y')
                actual_end = max(dates).strftime('%b %d, %Y')
                period_text = f"{actual_start} – {actual_end}"
            else:
                period_text = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
        except Exception:
            period_text = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
    else:
        period_text = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"

    # Document with custom margins
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=0.8 * inch,
        rightMargin=0.8 * inch,
        topMargin=1.1 * inch,
        bottomMargin=0.8 * inch
    )

    styles = getSampleStyleSheet()
    # Core styles
    h_style = ParagraphStyle(
        'HeadingMain', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=18,
        textColor=colors.HexColor('#1f2937'), spaceAfter=4, leading=22)
    sub_style = ParagraphStyle(
        'SubHeading', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#4b5563'),
        spaceAfter=16, leading=14, alignment=0)
    section_style = ParagraphStyle(
        'Section', parent=styles['Heading2'], fontSize=13, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#111827'), spaceBefore=18, spaceAfter=10)
    small_label = ParagraphStyle(
        'SmallLabel', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#6b7280'), leading=10)
    empty_style = ParagraphStyle(
        'Empty', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#6b7280'),
        alignment=1, spaceBefore=30)

    story = []

    # Title block (content - header graphics drawn in canvas)
    # Resolve office & campus for contextual header details
    office = Office.query.get(current_user.office_admin.office_id) if current_user and getattr(current_user, 'office_admin', None) else None
    campus_name = office.campus.name if office and office.campus else 'Campus'
    office_name = office.name if office else 'Office'
    story.append(Paragraph(f"{report_type.title()} Report", h_style))
    story.append(Paragraph(f"<b>{campus_name} • {office_name}</b><br/>Reporting Period: <b>{period_text}</b><br/>Generated: {generated_at_text}", sub_style))

    # --- SUMMARY / METRICS SECTION ---
    if report_type == 'inquiries':
        story.append(Paragraph('Summary Overview', section_style))
        total = len(data)
        # Status counts
        status_counts = {}
        for i in data:
            st = i.status.replace('_', ' ').title()
            status_counts[st] = status_counts.get(st, 0) + 1

        # Build status row (cards)
        status_row_headers = []
        status_row_values = []
        palette = ['#1d4ed8', '#059669', '#d97706', '#dc2626', '#6366f1']
        for idx, (st, count) in enumerate(sorted(status_counts.items())):
            color = colors.HexColor(palette[idx % len(palette)])
            status_row_headers.append(Paragraph(f"<font color='{color}'>■</font> {st}", styles['Normal']))
            status_row_values.append(Paragraph(f"<b>{count}</b>", styles['Normal']))

        summary_data = [
            [Paragraph('<b>Total Inquiries</b>', styles['Normal']), Paragraph(f"<b>{total}</b>", styles['Normal'])]
        ]
        if status_row_headers:
            summary_data.append(status_row_headers)
            summary_data.append(status_row_values)

        summary_table = Table(summary_data, hAlign='LEFT')
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ffffff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
            ('LINEBELOW', (0, 0), (-1, 0), 0.6, colors.HexColor('#e5e7eb')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
        ]))
        story.append(summary_table)

    elif report_type == 'counseling':
        story.append(Paragraph('Summary Overview', section_style))
        total = len(data)
        video_count = len([s for s in data if s.is_video_session])
        in_person_count = total - video_count
        status_counts = {}
        for s in data:
            st = s.status.replace('_', ' ').title()
            status_counts[st] = status_counts.get(st, 0) + 1
        status_cells = []
        for st, count in sorted(status_counts.items()):
            status_cells.append(Paragraph(f"<b>{count}</b><br/><font size=8 color='#6b7280'>{st}</font>", styles['Normal']))
        summary_data = [
            [Paragraph('<b>Total Sessions</b>', styles['Normal']), Paragraph(f"<b>{total}</b>", styles['Normal']),
             Paragraph('<b>Video</b>', styles['Normal']), Paragraph(str(video_count), styles['Normal']),
             Paragraph('<b>In-Person</b>', styles['Normal']), Paragraph(str(in_person_count), styles['Normal'])]
        ]
        if status_cells:
            summary_data.append(status_cells)
        summary_table = Table(summary_data, hAlign='LEFT')
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
            ('LINEBELOW', (0, 0), (-1, 0), 0.6, colors.HexColor('#e5e7eb')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
        ]))
        story.append(summary_table)

    # Detailed table section
    story.append(Paragraph('Detailed Records', section_style))

    if not data:
        story.append(Paragraph('No records found for the selected criteria.', empty_style))
    else:
        if report_type == 'inquiries':
            table_data = [['ID', 'Subject', 'Student', 'Status', 'Date', 'Concerns']]
            status_colors = {}
            for inquiry in data:
                concerns = ', '.join([c.concern_type.name for c in inquiry.concerns])
                subject_text = inquiry.subject if len(inquiry.subject) <= 60 else inquiry.subject[:57] + '…'
                concerns_text = concerns if len(concerns) <= 55 else concerns[:52] + '…'
                status = inquiry.status.replace('_', ' ').title()
                table_data.append([
                    str(inquiry.id),
                    subject_text,
                    inquiry.student.user.get_full_name(),
                    status,
                    inquiry.created_at.strftime('%m/%d/%Y'),
                    concerns_text
                ])
                if status not in status_colors:
                    # Assign stable color
                    mapping = {
                        'Pending': '#f59e0b', 'In Progress': '#3b82f6', 'Resolved': '#10b981',
                        'Closed': '#6b7280'
                    }
                    status_colors[status] = mapping.get(status, '#6366f1')
            col_widths = [0.6*inch, 2.4*inch, 1.6*inch, 1.0*inch, 0.9*inch, 1.6*inch]
        else:  # counseling
            table_data = [['ID', 'Student', 'Status', 'Scheduled', 'Duration', 'Type']]
            status_colors = {}
            for session in data:
                status = session.status.replace('_', ' ').title()
                table_data.append([
                    str(session.id),
                    session.student.user.get_full_name(),
                    status,
                    session.scheduled_at.strftime('%m/%d/%Y %I:%M %p') if session.scheduled_at else '—',
                    f"{session.duration_minutes}m" if session.duration_minutes else '—',
                    'Video' if session.is_video_session else 'In-Person'
                ])
                if status not in status_colors:
                    mapping = {
                        'Scheduled': '#3b82f6', 'Completed': '#10b981', 'Cancelled': '#dc2626',
                        'In Progress': '#6366f1'
                    }
                    status_colors[status] = mapping.get(status, '#6b7280')
            col_widths = [0.5*inch, 2.1*inch, 1.1*inch, 1.6*inch, 0.8*inch, 1.0*inch]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        # Base style
        table_style = [
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 9.5),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, 0), 7), ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            ('FONTSIZE', (0, 1), (-1, -1), 8.5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e5e7eb')),
            ('LEFTPADDING', (0, 0), (-1, -1), 5), ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]
        # Center numeric/id columns
        table_style += [('ALIGN', (0, 1), (0, -1), 'CENTER')]
        # Status column shading
        status_col_index = 3 if report_type == 'inquiries' else 2
        for row_idx in range(1, len(table_data)):
            status_value = table_data[row_idx][status_col_index]
            hex_color = status_colors.get(status_value, '#6b7280')
            table_style.append(('BACKGROUND', (status_col_index, row_idx), (status_col_index, row_idx), colors.HexColor(hex_color)))
            table_style.append(('TEXTCOLOR', (status_col_index, row_idx), (status_col_index, row_idx), colors.white))
            table_style.append(('FONTNAME', (status_col_index, row_idx), (status_col_index, row_idx), 'Helvetica-Bold'))

        table.setStyle(TableStyle(table_style))
        story.append(table)

        # Legend
        legend_parts = []
        for st, clr in status_colors.items():
            legend_parts.append(f"<font color='{clr}'>■</font> {st}")
        if legend_parts:
            legend_para = Paragraph('Status Legend: ' + ' &nbsp; '.join(legend_parts), small_label)
            story.append(Spacer(1, 10))
            story.append(legend_para)

    # Canvas header/footer callbacks with professional logo placement
    def _header_footer(canvas, doc):
        canvas.saveState()
        header_height = 50
        canvas.setFillColor(colors.HexColor('#111827'))
        canvas.rect(0, A4[1] - header_height, A4[0], header_height, stroke=0, fill=1)

        # Resolve office early (used for both text and potential branded logo)
        office = Office.query.get(current_user.office_admin.office_id) if current_user and getattr(current_user, 'office_admin', None) else None
        campus_name = office.campus.name if office and office.campus else ''
        office_name = office.name if office else ''

        # Determine logo path (office-specific fallback chain)
        logo_path = get_logo_path(office)
        text_x = doc.leftMargin
        if logo_path:
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                max_h = 38
                max_w = 120
                scale = min(max_w / iw, max_h / ih)
                dw, dh = iw * scale, ih * scale
                # Center vertically in header bar
                y = A4[1] - header_height + (header_height - dh) / 2
                canvas.drawImage(logo_path, doc.leftMargin, y, width=dw, height=dh, preserveAspectRatio=True, mask='auto')
                text_x = doc.leftMargin + dw + 12
            except Exception:
                # On failure, keep default text_x
                pass
        else:
            # If no logo, nudge text slightly for visual balance
            text_x = doc.leftMargin + 2

        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 12)
        canvas.drawString(text_x, A4[1] - 28, 'PiyuGuide Reports')
        canvas.setFont('Helvetica', 8)
        canvas.drawString(text_x, A4[1] - 40, f'{campus_name} • {office_name} • {report_type.title()} • {period_text}')

        # Footer (page metadata)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.setFont('Helvetica', 7)
        page_num = canvas.getPageNumber()
        footer_text = f"Generated {generated_at_text} • Page {page_num}"
        canvas.drawString(doc.leftMargin, 0.55 * inch, footer_text)
        canvas.drawRightString(A4[0] - doc.rightMargin, 0.55 * inch, 'Confidential')
        canvas.restoreState()

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={report_type}_report.pdf'}
    )


def generate_summary_pdf(stats, office_id, start_date, end_date):
    """Generate enhanced summary PDF report with unified styling (header/footer, metric cards, insights), brand rename, and campus/office context."""
    output = io.BytesIO()
    generated_at = datetime.now()
    generated_at_text = generated_at.strftime('%B %d, %Y • %I:%M %p')

    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=0.85 * inch,
        rightMargin=0.85 * inch,
        topMargin=1.2 * inch,
        bottomMargin=0.85 * inch
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=20, fontName='Helvetica-Bold',
                                 textColor=colors.HexColor('#111827'), spaceAfter=6, leading=24)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#4b5563'),
                               spaceAfter=18, leading=14)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#111827'), spaceBefore=16, spaceAfter=12)
    insight_style = ParagraphStyle('Insight', parent=styles['Normal'], fontSize=10.5, leading=14,
                                   textColor=colors.HexColor('#1f2937'), spaceBefore=6, spaceAfter=6)
    small_muted = ParagraphStyle('Muted', parent=styles['Normal'], fontSize=7.5, textColor=colors.HexColor('#6b7280'))

    story = []
    office = Office.query.get(office_id)
    campus_name = office.campus.name if office and office.campus else 'Campus'
    office_name = office.name if office else 'Office'
    period_text = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
    story.append(Paragraph('Monthly Summary Report', title_style))
    story.append(Paragraph(f'<b>{campus_name} • {office_name}</b><br/>Period: {period_text}<br/>Generated: {generated_at_text}', sub_style))

    # EXEC SUMMARY
    story.append(Paragraph('Executive Overview', section_style))
    total_interactions = stats['total_inquiries'] + stats['total_sessions']
    monthly_total = stats['inquiries_this_month'] + stats['sessions_this_month']
    if total_interactions:
        activity_ratio = (monthly_total / total_interactions) * 100
    else:
        activity_ratio = 0
    exec_data = [
        ['Total Interactions', f"{total_interactions:,}"],
        ['Monthly Activity Share', f"{activity_ratio:.1f}%"],
        ['Inquiry Response Rate', f"{stats['response_rate']}%"],
        ['Avg Resolution Time', stats['avg_resolution_time']]
    ]
    exec_table = Table(exec_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch, 2.2*inch])
    exec_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e5e7eb')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8)
    ]))
    story.append(exec_table)

    # Metrics split
    story.append(Paragraph('Core Metrics', section_style))
    metrics_data = [
        ['Metric', 'Value', 'Metric', 'Value'],
        ['Total Inquiries', f"{stats['total_inquiries']:,}", 'Total Sessions', f"{stats['total_sessions']:,}"],
        ['Inquiries (This Month)', f"{stats['inquiries_this_month']:,}", 'Sessions (This Month)', f"{stats['sessions_this_month']:,}"],
        ['Response Rate', f"{stats['response_rate']}%", 'Avg Resolution Time', stats['avg_resolution_time']]
    ]
    metrics_table = Table(metrics_data, colWidths=[1.9*inch, 1.3*inch, 1.9*inch, 1.3*inch], hAlign='LEFT')
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'), ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f9fafb')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8)
    ]))
    story.append(metrics_table)

    # Insights
    story.append(Paragraph('Insights & Recommendations', section_style))
    insights = []
    rr = stats['response_rate']
    if rr >= 80:
        insights.append('Strong inquiry management efficiency maintained (≥80% response).')
    elif rr >= 60:
        insights.append('Response performance moderate; consider process tuning to reach 80%.')
    else:
        insights.append('Low response rate; prioritize workflow optimization and staffing review.')
    if stats['sessions_this_month'] > stats['inquiries_this_month']:
        insights.append('Counseling engagement outpaces inquiry submissions—leverage success stories.')
    else:
        insights.append('Inquiry volume drives engagement—consider proactive outreach to convert inquiries to sessions.')
    if monthly_total > 50:
        insights.append('High monthly interaction volume indicates sustained demand; monitor capacity.')
    elif monthly_total < 15:
        insights.append('Low interaction volume—evaluate awareness campaigns or channel accessibility.')
    for text in insights:
        story.append(Paragraph(f'• {text}', insight_style))

    story.append(Spacer(1, 18))
    story.append(Paragraph('This report is generated automatically and intended for internal strategic planning.', small_muted))

    def _header_footer(canvas, doc):
        canvas.saveState()
        header_height = 50
        canvas.setFillColor(colors.HexColor('#111827'))
        canvas.rect(0, A4[1] - header_height, A4[0], header_height, stroke=0, fill=1)

        office_local = Office.query.get(office_id)
        campus_name = office_local.campus.name if office_local and office_local.campus else ''
        office_name = office_local.name if office_local else ''

        logo_path = get_logo_path(office_local)
        text_x = doc.leftMargin
        if logo_path:
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                max_h = 38
                max_w = 120
                scale = min(max_w / iw, max_h / ih)
                dw, dh = iw * scale, ih * scale
                y = A4[1] - header_height + (header_height - dh) / 2
                canvas.drawImage(logo_path, doc.leftMargin, y, width=dw, height=dh, preserveAspectRatio=True, mask='auto')
                text_x = doc.leftMargin + dw + 12
            except Exception:
                pass
        else:
            text_x = doc.leftMargin + 2

        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 12)
        canvas.drawString(text_x, A4[1] - 28, 'PiyuGuide Summary')
        canvas.setFont('Helvetica', 8)
        canvas.drawString(text_x, A4[1] - 40, f'{campus_name} • {office_name} • Monthly Performance Overview')

        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.setFont('Helvetica', 7)
        page_num = canvas.getPageNumber()
        canvas.drawString(doc.leftMargin, 0.55 * inch, f'Generated {generated_at_text} • Page {page_num}')
        canvas.drawRightString(A4[0] - doc.rightMargin, 0.55 * inch, 'Confidential')
        canvas.restoreState()

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment; filename=office_summary_report.pdf'}
    )


def generate_summary_excel(stats):
    """Generate summary Excel report"""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary Report"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    worksheet.cell(row=1, column=1, value="Metric").font = header_font
    worksheet.cell(row=1, column=1).fill = header_fill
    worksheet.cell(row=1, column=2, value="Value").font = header_font
    worksheet.cell(row=1, column=2).fill = header_fill
    
    # Data
    metrics = [
        ('Total Inquiries', stats['total_inquiries']),
        ('Inquiries This Month', stats['inquiries_this_month']),
        ('Total Counseling Sessions', stats['total_sessions']),
        ('Sessions This Month', stats['sessions_this_month']),
        ('Response Rate', f"{stats['response_rate']}%"),
        ('Average Resolution Time', stats['avg_resolution_time'])
    ]
    
    for row, (metric, value) in enumerate(metrics, 2):
        worksheet.cell(row=row, column=1, value=metric)
        worksheet.cell(row=row, column=2, value=value)
    
    # Auto-adjust column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 20
    
    workbook.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=summary_report.xlsx'}
    )


def get_status_breakdown_text(inquiries):
    """Get status breakdown text for inquiries"""
    status_counts = {}
    for inquiry in inquiries:
        status = inquiry.status.replace('_', ' ').title()
        status_counts[status] = status_counts.get(status, 0) + 1
    
    breakdown_parts = []
    for status, count in status_counts.items():
        breakdown_parts.append(f"{status}: {count}")
    
    return " | ".join(breakdown_parts) if breakdown_parts else "No data"


def get_session_status_breakdown_text(sessions):
    """Get status breakdown text for counseling sessions"""
    status_counts = {}
    for session in sessions:
        status = session.status.replace('_', ' ').title()
        status_counts[status] = status_counts.get(status, 0) + 1
    
    breakdown_parts = []
    for status, count in status_counts.items():
        breakdown_parts.append(f"{status}: {count}")
    
    return " | ".join(breakdown_parts) if breakdown_parts else "No data"
